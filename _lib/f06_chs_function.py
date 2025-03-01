"""
VENTANA PARA LA FUNCIÓN CHS
"""
from customtkinter import * 
import tkinter.messagebox
import openpyxl
import shutil
from _lib.get_files_path import get_file_paths
import xlwings as xw
from _lib.kill_excel_process import kill_excel_processes
from _lib.parse_coordinates import parse_coordinates
import time
import math
import PyPDF2
from _lib.progress_window import create_progress_window
from _lib.set_cell_border import set_border

def open_chs_window(menu_window,images_path):

    # Crear una nueva ventana
    window = CTkToplevel(menu_window)
    # Geometría
    width =410
    height = 200
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window.geometry(f'{width}x{height}+{x}+{y}')
    # Nombre de la ventana
    window.title("CHS")
    # Ícono ventana
    window.after(201, lambda: window.iconbitmap(os.path.join(images_path, "chs.ico")))
    # Resizable
    window.resizable(False, False)
    window.grab_set()  # Esto hace que la ventana sea modal
    window.transient(menu_window)  # Para que esté asociada a la ventana principal
    window.focus_force()  # Enfoca la ventana modal

    #Variables para almacenar archivos
    file_content = {"cliente":"",
                    "proyecto": "",
                    "orden_servicio": "",
                    "operador": "",
                    "interpreto": "",
                    "fecha_medicion": "",
                    "prof_ensayo": "",
                    "nombre_ensayo": "",
                    "data": {"depth":[],
                             "vs_internal":[],
                             "dist_fr" : [],
                             "t_llegada_onda" : []
                             }
                    }
    input_path = ""

    #Función para adjuntar archivo del input
    def open_file_dialog(browse_entry):

        nonlocal file_content

        #ruta del archivo
        file_path = filedialog.askopenfilename()
        #extensión del archivo
        file_extension = os.path.splitext(file_path)[1]
        #nombre del archivo
        file_name = os.path.basename(file_path)  # Obtiene solo el nombre del archivo

        if file_extension != ".xlsx":
            tkinter.messagebox.showerror("Error","Formato de archivo no válido.")
            return
        else:
            browse_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
            browse_entry.delete(0, "end")  # Limpia el contenido actual del campo
            browse_entry.insert(0, file_name)  # Inserta la ruta del archivo en el campo
            browse_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite

            #Abre el archivo
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            #Verifica que exista la información del proyecto, cliente y OS
            if sheet['B1'].value != None and sheet['B2'].value != None and sheet['B3'].value != None and sheet['B4'].value != None and sheet['B5'].value != None and sheet['B6'].value != None and sheet['B7'].value != None and sheet['B8'].value != None:
                file_content["proyecto"] = sheet['B1'].value
                file_content["cliente"] = sheet['B2'].value
                file_content["orden_servicio"] = sheet['B3'].value
                file_content["operador"] = sheet['B4'].value
                file_content["interpreto"] = sheet['B5'].value
                file_content["fecha_medicion"] = sheet['B6'].value
                file_content["prof_ensayo"] = sheet['B7'].value
                file_content["nombre_ensayo"] = sheet['B8'].value

            else:
                tkinter.messagebox.showerror("Error", "El archivo está vacío o no está en el formato correcto.")
                file_content = {"cliente":"",
                    "proyecto": "",
                    "orden_servicio": "",
                    "operador": "",
                    "interpreto": "",
                    "fecha_medicion": "",
                    "prof_ensayo": "",
                    "nombre_ensayo": "",
                    "data": {"depth":[],
                             "vs_internal":[],
                             "dist_fr" : [],
                             "t_llegada_onda" : []
                             }
                }
                browse_entry.configure(state=NORMAL)
                browse_entry.delete(0, "end")
                browse_entry.configure(state=DISABLED) 
                
                return
        
            #Última fila con información
            ultima_fila = 17
            for fila in range(17, sheet.max_row + 1):
                celda = sheet[f'A{fila}']
                if celda.value:  # Verifica si la celda tiene contenido
                    ultima_fila = fila

            if ultima_fila == 17:
                tkinter.messagebox.showerror("Error", "El archivo está vacío o no está en el formato correcto.")
                file_content = {"cliente":"",
                    "proyecto": "",
                    "orden_servicio": "",
                    "operador": "",
                    "interpreto": "",
                    "fecha_medicion": "",
                    "prof_ensayo": "",
                    "nombre_ensayo": "",
                    "data": {"depth":[],
                             "vs_internal":[],
                             "dist_fr" : [],
                             "t_llegada_onda" : []
                             }
                }
                browse_entry.configure(state=NORMAL)
                browse_entry.delete(0, "end")
                browse_entry.configure(state=DISABLED) 
                
                return
            
            #Obtiene los documentos a generar con sus respectivas coordenadas
            for row in sheet.iter_rows(min_row=18, max_row=ultima_fila, values_only=True):  # Ajusta según la tabla
                
                file_content["data"]["depth"].append(row[0])
                file_content["data"]["vs_internal"].append(row[3])
                file_content["data"]["dist_fr"].append(row[21])
                file_content["data"]["t_llegada_onda"].append(row[1] / 100 )

    #Función para adjuntar archivo del input
    def open_file_path(directory_entry, file_content):

        nonlocal input_path

        #ruta del archivo
        file_path = filedialog.askdirectory(title="Seleccionar un directorio")
        try:
            #contenido del directorio
            directory_content = os.listdir(file_path)
        except:
            tkinter.messagebox.showerror("Error", "La ruta no puede estar vacía.")
            return

        #Archivos necesarios para completar la verificación
        verify_files = ['foto.jpg','localizacion.jpg','registro.png']


        #Verifica que se haya insertado el input
        if file_content["cliente"] != "" and file_content["proyecto"] != "" and file_content["orden_servicio"] != "" and file_content["operador"] != "" and file_content["interpreto"] != "" and file_content["fecha_medicion"] != "" and file_content["prof_ensayo"] != "" and file_content["nombre_ensayo"] != "":
            
            #Verifica que en el directorio existan las carpetas correspondientes a los archivos
            try:
                assert set(os.listdir(f'{file_path}//')) == set(verify_files), "Los elementos no coinciden"

                directory_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
                directory_entry.delete(0, "end")  # Limpia el contenido actual del campo
                directory_entry.insert(0, file_path)  # Inserta la ruta del archivo en el campo
                directory_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite

            except AssertionError as e:

                directory_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
                directory_entry.delete(0, "end")  # Limpia el contenido actual del campo
                directory_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite

                tkinter.messagebox.showerror("Error", "No están todos los archivos necesarios para ejecutar el aplicativo.")
                return
        
        else:
            tkinter.messagebox.showerror("Error", "Debe seleccionar primero el archivo de entrada.")
            directory_entry.configure(state=NORMAL)  # Habilita el campo para insertar el texto
            directory_entry.delete(0, "end")  # Limpia el contenido actual del campo
            directory_entry.configure(state=DISABLED)  # Desactiva nuevamente el campo para que no se edite
            return

    #Label seleccionar archivo de entrada
    label_archivo_entrada = CTkLabel(master=window,text="Archivo de entrada:",fg_color="transparent",font=('Gothic A1',15))
    label_archivo_entrada.place(x= 20,y=20)

    #Cajita para almacenar el nombre del archivo
    browse_entry = CTkEntry(master=window, width=140, font=('Gothic A1', 12), placeholder_text="Archivo",state=DISABLED)
    browse_entry.place(x=160, y=20)

    # Botón para buscar archivo
    browse_button = CTkButton(
        master=window,
        text="Examinar",
        width=80,
        height=22,
        command= lambda: open_file_dialog(browse_entry)  
    )
    browse_button.place(x=310, y=23)

    #Label seleccionar archivos complementarios del input
    label_directory_input = CTkLabel(master=window,text="Directorio de entrada:",fg_color="transparent",font=('Gothic A1',15))
    label_directory_input.place(x= 20,y=80)

    #Cajita para almacenar la ruta
    directory_entry = CTkEntry(master=window, width=125, font=('Gothic A1', 12), placeholder_text="Directorio",state=DISABLED)
    directory_entry.place(x=175, y=80)

    # Botón para buscar ruta
    browse_path_button = CTkButton(
        master=window,
        text="Examinar",
        width=80,
        height=22,
        command= lambda: open_file_path(directory_entry,file_content)  
    )
    browse_path_button.place(x=310, y=83)

    # Ejecutar 
    run_button = CTkButton(
        master=window,
        text="Generar reportes",
        width=100,
        height=30,
        command= lambda: chs_module(file_content,directory_entry.get(),window) 
    )
    run_button.place(x=(width - 100) // 2, y=140)
    

"""
EJECUTA EL MÓDULO DE CHS
"""
def chs_module(file_content,inputs_path,window):

    #Verifica que se haya seleccionado la ruta de lso archivos de entrada
    if inputs_path == "":
        tkinter.messagebox.showerror("Error", "Favor seleccionar la ruta de los archivos de entrada.")
        return
    
    #Verifica que se haya seleccionado el archivo de entrada
    if file_content["cliente"] == "" and file_content["proyecto"] == "" and file_content["orden_servicio"] == "" and file_content["operador"] == "" and file_content["interpreto"] == "" and file_content["fecha_medicion"] == "" and file_content["prof_ensayo"] == "" and file_content["nombre_ensayo"] == "":
        tkinter.messagebox.showerror("Error", "Favor cargar archivo de entrada.")
        return

    #Cierra todos los procesos de excel abiertos
    kill_excel_processes()
    time.sleep(2)

    #Templates path
    templates_path = get_file_paths("_templates")

    #Ruta de documentos
    documents_path = os.path.expanduser("~\\Documents")

    # Crea la carpeta para almacenar
    results_path = f'{documents_path}//GEOSTREAM//CHS'
    if not os.path.exists(results_path):
        os.makedirs(results_path)
    else:
        # Elimina los archivos adentro si la carpeta ya existe
        for elemento in os.listdir(results_path):
            ruta_elemento = os.path.join(results_path, elemento)
            if os.path.isdir(ruta_elemento):
                shutil.rmtree(ruta_elemento)  
            else:
                os.remove(ruta_elemento)

    #Crea la ventana de progreso
    ventana_progreso, barra_progreso, texto_progreso = create_progress_window("CHS - Generando memorias","logo.ico",f"0/1")

    #Copia el template
    shutil.copy(f'{templates_path}//template chs.xlsm', f'{documents_path}///GEOSTREAM//CHS//template chs.xlsm')

    #Abre el template
    app = xw.App(visible=False)  
    app.display_alerts = False 
    app.screen_updating = False
    libro = xw.Book(f'{documents_path}//GEOSTREAM//CHS//template chs.xlsm', update_links=True)

    #Modifica header
    modificar_header(libro,file_content)

    #Modifica tabla
    modificar_resultados(libro,file_content)

    #Modificar hoja fotos
    modificar_hoja_fotos(libro,inputs_path)

    #Imprimir en PDF la memoria
    save_to_pdf(libro, documents_path )

    # Guardar los cambios
    libro.save()

    #Cerrar
    libro.close()
    app.quit()

    #Actualizar barra de progreso
    barra_progreso.set(1 / 1)  # Actualizar progreso
    texto_progreso.set(f"1/1")
    ventana_progreso.update_idletasks()  # Forzar actualización de la ventana

    ventana_progreso.destroy()
    tkinter.messagebox.showinfo("Info","Memorias generadas. Revisar en Documentos.")
    window.destroy()

def modificar_header(libro,file_content):

    hoja = libro.sheets["Resultados"] 
    hoja["C7"].value = file_content["proyecto"]
    hoja["C8"].value = file_content["cliente"]
    hoja["D9"].value = file_content["orden_servicio"]
    hoja["C10"].value = file_content["operador"]
    hoja["C11"].value = file_content["interpreto"]
    hoja["L7"].value = file_content["fecha_medicion"]
    hoja["L8"].value = file_content["prof_ensayo"]

def modificar_resultados(libro,file_content):

    hoja = libro.sheets["Resultados"] 

    row_inicio = 15
    for i in range(len(file_content["data"]["depth"])):
        hoja[f"B{row_inicio + i}"].value = file_content["data"]["depth"][i]
        hoja[f"C{row_inicio + i}"].value = file_content["data"]["dist_fr"][i]
        hoja[f"D{row_inicio + i}"].value = file_content["data"]["t_llegada_onda"][i]
        hoja[f"F{row_inicio + i}"].value = file_content["data"]["vs_internal"][i]


def modificar_hoja_fotos(libro,inputs_path):

    #Parsea el input_path
    inputs_path = inputs_path.replace("/", "\\")

    #Obtiene la hoja
    hoja = libro.sheets["Resultados"]

    #Borra y pega la imagen
    left_inicial = hoja.range("B114").left
    desplazamiento_horizontal = hoja.range("B114").width / 2
    hoja.pictures.add(f"{inputs_path}\\registro.png",
                    top = hoja.range("B114").top,
                    left=left_inicial + desplazamiento_horizontal,
                    width = 900,
                    height=350)
    
    #Borra y pega la imagen
    left_inicial = hoja.range("B140").left
    desplazamiento_horizontal = hoja.range("B140").width / 2
    hoja.pictures.add(f"{inputs_path}\\foto.jpg",
                    top=hoja.range("B140").top,
                    left=left_inicial + desplazamiento_horizontal,
                    width = 300,
                    height=400)
    
    #Borra y pega la imagen
    hoja.pictures.add(f"{inputs_path}\\localizacion.jpg",
                    top=hoja.range("F140").top,
                    left=hoja.range("F140").left,
                    width = 580,
                    height=400)
    
   
def save_to_pdf(libro, documents_path ):

    #hojas
    hoja = libro.sheets["Resultados"]

    #Guarda el pdf
    output_pdf_path = f'{documents_path}//GEOSTREAM//CHS//combinado.pdf'

    hoja.api.ExportAsFixedFormat(
        Type=0,  # 0 es para PDF
        Filename=output_pdf_path,
        Quality=0,  # Calidad estándar
        IncludeDocProperties=True,
        IgnorePrintAreas=False,  # Respetar áreas de impresión
        OpenAfterPublish=False
    )   

